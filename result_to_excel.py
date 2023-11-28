# -----------------------------------------------------------------------------#
# Script Name: result_to_excel.py
# Description: Export results to an Excel file while applying modifications.
#
# Usage example:
#   1. Read relevant test results from CSV files.
#   2. Apply modifications to the data, such as dividing columns by specific values.
#   3. Write the modified data to an existing Excel file.
#
# Requirements:
#   - Python packages: pandas, openpyxl
#
# Author: liuyupeng
# Date: November 2, 2023
# -----------------------------------------------------------------------------#


import pandas as pd
from openpyxl import load_workbook

def res_to_excel(mode, interface, csv_path, excl_path, sheet, start_row):
    # 读取CSV文件
    df = pd.read_csv(csv_path)

    # 选择感兴趣的列
    selected_columns = df[["Agg Tx Rate (% Line Rate)", "Agg Rx Throughput (Mbps)", "Avg Latency (ns)"]]

    # 创建 selected_columns 的深拷贝
    selected_columns_copy = selected_columns.copy()

    # 将 "Agg Tx Rate (% Line Rate)" 列除以 100
    selected_columns_copy["Agg Tx Rate (% Line Rate)"] = selected_columns_copy["Agg Tx Rate (% Line Rate)"] / 100

    # 将 "Avg Latency (ns)" 列除以 1000
    selected_columns_copy["Avg Latency (ns)"] = selected_columns_copy["Avg Latency (ns)"] / 1000

    # 打印修改后的DataFrame
    print(selected_columns_copy)

    # 打开现有的 Excel 文件
    book = load_workbook(excl_path)

    # 获取xlsx工作表
    ws = book[sheet]

    # 找到要写入数据的起始位置
    start_row, start_col = start_row, 8

    # 将数据写入工作表的指定位置，不影响其他数据和格式
    for index, row in selected_columns_copy.iterrows():
        for col_idx, col_name in enumerate(selected_columns_copy.columns, 0):
            ws.cell(row=start_row + index, column=start_col + col_idx, value=row[col_name])

    # 写入mode和interface
    ws.cell(row=start_row, column=start_col - 6, value=mode)
    ws.cell(row=start_row, column=start_col + 3, value=interface)

    # 保存更改
    book.save(excl_path)

def read_excl(result_path, sheet, start_row):
    df = pd.read_excel(result_path)
    selected = df[["测试配置", "本地保存路径", "接口信息", "模式"]]

    for index, row in selected.iterrows():
        csv_path = row["本地保存路径"]
        interface = row["接口信息"]
        mode = row["测试配置"] + "\n" + row["模式"]
        res_to_excel(mode, interface, csv_path, excl_path, sheet, start_row)
        start_row += 7

        # 在这里你可以使用上述提取到的参数进行操作
        print(f"本地保存路径: {csv_path}")
        print(f"测试配置: {mode}")
        print(f"接口信息: {interface}")

excl_path = "xlsx/FW-RFC2544.xlsx"
sheet = "V2800"
start_row = 8
result_path = 'result/2023-11-01_20-12-53_result.xlsx'
read_excl(result_path, sheet, start_row)
