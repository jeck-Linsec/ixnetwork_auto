# -----------------------------------------------------------------------------#
# Script Name: ixnetwork_test.py
# Description: Import related configurations for testing.
#
# Usage example:
#   1. Import basic information such as the tester IP address and serial port
#   2. Import the test device and configuration list
#   3. Iterate through the test and log until you have traversed the entire list
#
# Requirements:
#   -
#
# Author: liuyupeng
# Date: November 2, 2023
# -----------------------------------------------------------------------------#

import logging
import yaml
import datetime
import time
import firewall_config
import rfc2544_test
from colorama import Fore, init
from openpyxl import load_workbook
import shutil

# # 打开现有的 Excel 文件
# excel_file = "result/result.xlsx"
# wb = load_workbook(excel_file)
# ws = wb.worksheets[0]

# 初始化 colorama
init(autoreset=True)

def print_cool_logo():

#  ______  __   __   __  __  ____    ______      ______  ____    ____    ______
# /\__  _\/\ \ /\ \ /\ \/\ \/\  _`\ /\__  _\    /\__  _\/\  _`\ /\  _`\ /\__  _\
# \/_/\ \/\ `\`\/'/'\ \ `\\ \ \ \L\_\/_/\ \/    \/_/\ \/\ \ \L\_\ \,\L\_\/_/\ \/
#    \ \ \ `\/ > <   \ \ , ` \ \  _\L  \ \ \       \ \ \ \ \  _\L\/_\__ \  \ \ \
#     \_\ \__ \/'/\`\ \ \ \`\ \ \ \L\ \ \ \ \       \ \ \ \ \ \L\ \/\ \L\ \ \ \ \
#     /\_____\/\_\\ \_\\ \_\ \_\ \____/  \ \_\       \ \_\ \ \____/\ `\____\ \ \_\
#     \/_____/\/_/ \/_/ \/_/\/_/\/___/    \/_/        \/_/  \/___/  \/_____/  \/_/
    print(Fore.YELLOW + r"""

╔────────────────────────────────────────────────────────────────────────────────────────────────────────────────╗
│   ____    ____    ____        ___    ______  __ __    __ __            ______  ____    ____    ______  __      │
│  /\  _`\ /\  _`\ /\  _`\    /'___`\ /\  ___\/\ \\ \  /\ \\ \          /\__  _\/\  _`\ /\  _`\ /\__  _\/\ \     │
│  \ \ \L\ \ \ \L\_\ \ \/\_\ /\_\ /\ \\ \ \__/\ \ \\ \ \ \ \\ \         \/_/\ \/\ \ \L\_\ \,\L\_\/_/\ \/\ \ \    │
│   \ \ ,  /\ \  _\/\ \ \/_/_\/_/// /__\ \___``\ \ \\ \_\ \ \\ \_   _______\ \ \ \ \  _\L\/_\__ \  \ \ \ \ \ \   │
│    \ \ \\ \\ \ \/  \ \ \L\ \  // /_\ \\/\ \L\ \ \__ ,__\ \__ ,__\/\______\\ \ \ \ \ \L\ \/\ \L\ \ \ \ \ \ \_\  │
│     \ \_\ \_\ \_\   \ \____/ /\______/ \ \____/\/_/\_\_/\/_/\_\_/\/______/ \ \_\ \ \____/\ `\____\ \ \_\ \/\_\ │
│      \/_/\/ /\/_/    \/___/  \/_____/   \/___/    \/_/     \/_/             \/_/  \/___/  \/_____/  \/_/  \/_/ │
╚────────────────────────────────────────────────────────────────────────────────────────────────────────────────╝
    """)


print_cool_logo()
# 定义测试开始和结束标识字符
TEST_START = "********************************************************\n" \
             "********************************                        开始测试                           *\n" \
             "********************************************************************************************"
TEST_END = "********************************************************\n" \
           "********************************                        结束测试                           *\n" \
           "********************************************************************************************"

# 获取当前的日期和时间
start_time = datetime.datetime.now()
current_time = start_time.strftime("%Y-%m-%d_%H-%M-%S")

# 定义模板文件和目标文件的文件名
template_file = 'result/template/template.xlsx'
new_file = f'result/{current_time}_result.xlsx'  # 新文件名

# 使用 shutil 复制模板文件到目标文件
shutil.copy(template_file, new_file)

# 打开 Excel 文件
excel_file = f"result/{current_time}_result.xlsx"
wb = load_workbook(excel_file)
ws = wb.worksheets[0]

# 配置日志记录
log_name = f'log/{current_time}.log'
logging.basicConfig(filename=log_name, level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

logging.info(TEST_START)
logging.info(f'开始时间:{start_time}')

# 读取配置文件
with open('config.yaml', 'r', encoding='utf-8') as file:
    config = yaml.safe_load(file)

logging.info("测试初始化，读取配置文件")


# BPS系统信息
ixNetworkApiServer = config['ixNetworkApiServer']
ixNetworkPort = config['ixNetworkPort']
quickTestConfigFiles = config['quickTestConfigFiles']
logging.info(f'ixNetwork测试仪信息:{ixNetworkApiServer} ,用户：{ixNetworkPort}, QT_File:{quickTestConfigFiles}\n')

# 获取设备串口信息
host = config['host']
port = config['port']
timeout = config['timeout']
# 获取设备1串口信息
host1 = config['host1']
port1 = config['port1']

# 要执行的命令列表
login = config['login']
config_file = config['config_file']
# config_Bridge = config['config_Bridge']
# config_gatway = config['config_gatway']
reset_session = config['reset_session']

print(quickTestConfigFiles)

for quickTestConfig in quickTestConfigFiles:
    interface = quickTestConfig['interface']
    interface_list = interface.split(',')
    mode = quickTestConfig['mode']
    key = None
    if mode == "Bridge":
        conf = firewall_config.create_bridge_config(interface_list)
    elif mode == "gateway":
        conf = firewall_config.create_gateway_config(interface_list)
    elif mode == "ipsec":
        key = quickTestConfig['key']
        conf = firewall_config.ipsec_config(key)
        # conf1 = conf.copy()
        # conf1[11] = 'active on'
        interface = ""
        firewall_config.configure_firew(host1, port1, conf, timeout, log_name)
    else:
        # 处理未知模式的情况
        print("Unknown mode:", mode)
        #conf = []
        break

    firewall_config.configure_firew(host, port, conf, timeout, log_name)
    if key:
        mode = mode + key
    time.sleep(30)

    print(mode)
    quickTestConfigFile = quickTestConfig['name']
    print(quickTestConfigFile)
    start_time1 = datetime.datetime.now()
    current_time1 = start_time1.strftime("%Y-%m-%d_%H-%M-%S")
    windowsPath, localPath = rfc2544_test.Rfc2544_run(ixNetworkApiServer, ixNetworkPort, quickTestConfigFile, log_name)

    # 生成测试数据
    test_data = [quickTestConfigFile, start_time1, windowsPath, localPath, interface, mode]
    ws.append(test_data)
    # 保存 Excel 文件
    wb.save(excel_file)
    time.sleep(5)
    # 清除会话
    firewall_config.configure_firew(host, port, reset_session, timeout, log_name)
    if mode == "ipsec":
        firewall_config.configure_firew(host1, port1, reset_session, timeout, log_name)
    # 删除配置
    if mode == "Bridge":
        conf = firewall_config.undo_bridge_config(interface_list)
        firewall_config.configure_firew(host, port, conf, timeout, log_name)
    if mode == "gateway":
        conf = firewall_config.undo_gateway_config(interface_list)
        firewall_config.configure_firew(host, port, conf, timeout, log_name)
    time.sleep(60)

# 获取当前的日期和时间
end_time = datetime.datetime.now()
logging.info(f'结束时间:{end_time}\n')
logging.info(TEST_END)